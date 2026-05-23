from __future__ import annotations

import hashlib
import io
import re
from typing import Any, BinaryIO, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..structs import LociMap, ParsedPlateLayout, WellLayout


DEFAULT_PLATE_FORMAT = "6x10"

RESERVED_COLS = {
    "well", "well_id", "id", "id #", "id#", "id no", "id no.",
    "id number", "id num", "ctrl", "control", "test", "combo_id",
    "hla_combination_id", "race",
}

ALLOWED_LOCI_CANON: Dict[str, str] = {
    "A": "A",
    "B": "B",
    "C": "C",
    "BW4": "Bw4",
    "BW6": "Bw6",
    "DR": "DR",
    "DRB1": "DRB1",
    "DRB3": "DRB3",
    "DRB4": "DRB4",
    "DRB5": "DRB5",
    "DQ": "DQ",
    "DQA1": "DQA1",
    "DQB1": "DQB1",
    "DP": "DP",
    "DPA1": "DPA1",
    "DPB1": "DPB1",
}

COMMON_PLATE_SIZES = {"6x10", "2x3", "4x6", "8x12", "16x24", "32x48"}


def _sha256_bytes(b: bytes) -> str:
    """Return the SHA-256 checksum of a byte string.

    Parameters
    ----------
    b : bytes
        Input bytes.

    Returns
    -------
    str
        Hex-encoded SHA-256 checksum.
    """
    return hashlib.sha256(b).hexdigest()


def _clean(s: Any) -> Optional[str]:
    """Convert a cell value to a stripped string or ``None``.

    Parameters
    ----------
    s : Any
        Input cell value.

    Returns
    -------
    str or None
        Stripped string value. Empty values return ``None``.
    """
    if s is None:
        return None

    s = str(s).strip()
    return s or None


def _normalize_well_id(val: str) -> Optional[str]:
    """Normalize a well ID.

    Parameters
    ----------
    val : str
        Well ID in either letter-number form, such as ``"A1"``, or
        number-letter form, such as ``"1A"``.

    Returns
    -------
    str or None
        Normalized well ID in letter-number form, or ``None`` if the value is
        not a valid well ID.
    """
    if not val:
        return None

    v = val.strip().upper()

    m = re.match(r"^([A-Z])([1-9][0-9]?)$", v)
    if m:
        return f"{m.group(1)}{m.group(2)}"

    m = re.match(r"^([1-9][0-9]?)([A-Z])$", v)
    if m:
        return f"{m.group(2)}{m.group(1)}"

    return None


def _normalize_locus_header(h: str) -> Optional[str]:
    """Normalize an Excel header to a known locus label.

    Parameters
    ----------
    h : str
        Header value from the Excel sheet.

    Returns
    -------
    str or None
        Canonical locus label if the header is allowed. Returns ``None`` for
        unknown headers, numeric headers, or non-locus columns.

    Notes
    -----
    Allele suffixes are removed before lookup. For example, ``"C*07:02"`` is
    reduced to ``"C"``.
    """
    if not h:
        return None

    x = h.strip()
    x = re.sub(r'["\s]+', "", x)
    x = x.rstrip(":")
    x = re.sub(r"\*.*$", "", x)

    if re.fullmatch(r"\d+", x):
        return None

    x_up = x.upper()
    x_up = {
        "DRB": "DRB1",
        "DQA": "DQA1",
        "DQB": "DQB1",
        "DPA": "DPA1",
        "DPB": "DPB1",
    }.get(x_up, x_up)

    return ALLOWED_LOCI_CANON.get(x_up)


def _infer_plate_format_from_wells(well_ids: List[str]) -> str:
    """Infer a plate format from well IDs.

    Parameters
    ----------
    well_ids : list of str
        Well IDs such as ``"A1"``, ``"H12"``, or ``"P24"``.

    Returns
    -------
    str
        Inferred plate format in ``"{rows}x{cols}"`` format. If no valid well
        IDs are present, returns ``DEFAULT_PLATE_FORMAT``.
    """
    if not well_ids:
        return DEFAULT_PLATE_FORMAT

    max_letter = "A"
    max_col = 1

    for wid in well_ids:
        wid = wid.strip().upper()
        m = re.match(r"^([A-Z])([1-9][0-9]?)$", wid)

        if not m:
            continue

        letter, num = m.group(1), int(m.group(2))

        if letter > max_letter:
            max_letter = letter

        if num > max_col:
            max_col = num

    n_rows = ord(max_letter) - ord("A") + 1
    return f"{n_rows}x{max_col}"


def parse_excel_layout(fileobj: BinaryIO) -> Tuple[str, ParsedPlateLayout]:
    """Parse a vendor Excel plate layout.

    Parameters
    ----------
    fileobj : BinaryIO
        Binary file-like object containing an Excel workbook.

    Returns
    -------
    sha256 : str
        SHA-256 checksum of the uploaded workbook bytes.
    layout : ParsedPlateLayout
        Parsed plate layout with normalized well IDs, metadata, loci, warnings,
        and inferred plate format.

    Raises
    ------
    ValueError
        If no valid header row can be found, or if the detected header row does
        not contain a well ID column.
    openpyxl.utils.exceptions.InvalidFileException
        If the workbook cannot be opened by ``openpyxl``.

    Notes
    -----
    The parser scans all sheets for lot and complement metadata, then parses
    the active sheet for the main well table. The first ID-like column is used
    as the well ID. Later ID-like columns are treated as HLA combination IDs.
    Only whitelisted locus columns are parsed.
    """
    raw = fileobj.read()
    sha256 = _sha256_bytes(raw)
    wb = load_workbook(io.BytesIO(raw), data_only=True)

    warnings: List[str] = []
    lot_no: Optional[str] = None
    compl_no: Optional[str] = None

    for ws in wb.worksheets:
        max_r = min(ws.max_row, 120)
        max_c = ws.max_column

        try:
            hf = ws.header_footer
            for text in (
                getattr(hf, "left_header", None),
                getattr(hf, "center_header", None),
                getattr(hf, "right_header", None),
                getattr(hf, "left_footer", None),
                getattr(hf, "center_footer", None),
                getattr(hf, "right_footer", None),
            ):
                if not text:
                    continue

                t = re.sub(r"&[^ ]*", "", str(text))
                low = t.lower()

                m = re.search(r"\blot[:\s]+([A-Za-z0-9\-_\/]+)", low)
                if m and not lot_no:
                    lot_no = m.group(1).upper()

                m = re.search(r"\bcomp(?:l|lete)?[:\s]+([A-Za-z0-9\-_\/]+)", low)
                if m and not compl_no:
                    compl_no = m.group(1).upper()
        except Exception:
            pass

        def _scan_right(row: int, col: int, span: int = 12) -> Optional[str]:
            """Return the first non-empty cell to the right of a cell.

            Parameters
            ----------
            row : int
                One-based row index.
            col : int
                One-based column index.
            span : int, optional
                Number of cells to inspect to the right. The default is ``12``.

            Returns
            -------
            str or None
                First non-empty value converted to uppercase, or ``None`` if
                no value is found.
            """
            for cc in range(col + 1, min(col + span + 1, max_c + 1)):
                v = _clean(ws.cell(row, cc).value)
                if v:
                    return v.strip().upper()

            return None

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                val = _clean(ws.cell(r, c).value)

                if not val:
                    continue

                low = val.lower()

                if "lot" in low and lot_no is None:
                    got = _scan_right(r, c, span=12)
                    if got:
                        lot_no = got

                if (
                    ("compl" in low)
                    or ("comp " in low)
                    or ("comp:" in low)
                    or ("complete" in low)
                ) and compl_no is None:
                    got = _scan_right(r, c, span=12)
                    if got:
                        compl_no = got

            if lot_no and compl_no:
                break

        if lot_no and compl_no:
            break

    ws = wb.active
    header_row_idx = None
    headers: List[str] = []

    for r in range(1, ws.max_row + 1):
        row_vals = [
            _clean(ws.cell(r, c).value)
            for c in range(1, ws.max_column + 1)
        ]

        while row_vals and row_vals[-1] is None:
            row_vals.pop()

        if not row_vals:
            continue

        lowered = [(v or "").lower() for v in row_vals]
        has_id = any(
            v in {"id", "id #", "id#", "id no", "id no.", "id number", "id num"}
            for v in lowered
        )
        has_race = any(v == "race" for v in lowered)
        banner = "well" in lowered and "reactions" in lowered

        if has_id and has_race:
            header_row_idx = r
            headers = [(h or "") for h in row_vals]
            break

        if banner:
            for rr in range(r + 1, min(r + 10, ws.max_row + 1)):
                row2 = [
                    _clean(ws.cell(rr, c).value)
                    for c in range(1, ws.max_column + 1)
                ]
                dense = [x for x in row2 if x]
                has_id_like = any(
                    (v or "").lower()
                    in {"id", "id #", "id#", "id no", "id no.", "id number", "id num"}
                    for v in row2
                )

                if len(dense) >= 5 and has_id_like:
                    header_row_idx = rr
                    headers = [(h or "") for h in row2]
                    break

            if header_row_idx is not None:
                break

    if header_row_idx is None:
        raise ValueError("Could not find a valid header row (need an 'ID' and 'Race' row).")

    norm_headers: List[str] = []
    norm_locus_headers: List[Optional[str]] = []
    id_assigned = False

    for h in headers:
        raw_header = (h or "").strip()
        low = raw_header.lower()

        if low in (
            "well", "well id", "well_id", "id", "id #", "id#",
            "id no", "id no.", "id number", "id num",
        ):
            if not id_assigned and low in ("well", "well id", "well_id", "id"):
                norm_headers.append("id")
                norm_locus_headers.append(None)
                id_assigned = True
            else:
                norm_headers.append("hla_combination_id")
                norm_locus_headers.append(None)
        elif low in ("race", "nationality", "donor_race"):
            norm_headers.append("race")
            norm_locus_headers.append(None)
        elif low in ("ctrl", "control"):
            norm_headers.append("ctrl")
            norm_locus_headers.append(None)
        elif low == "test":
            norm_headers.append("test")
            norm_locus_headers.append(None)
        else:
            locus = _normalize_locus_header(raw_header)
            norm_headers.append(raw_header)
            norm_locus_headers.append(locus)

    try:
        id_col_idx = norm_headers.index("id")
    except ValueError:
        raise ValueError("Header row does not contain a well ID column.")

    combo_col_idx = (
        norm_headers.index("hla_combination_id")
        if "hla_combination_id" in norm_headers
        else None
    )
    race_col_idx = norm_headers.index("race") if "race" in norm_headers else None

    locus_cols: List[Tuple[int, str]] = [
        (idx, canon)
        for idx, canon in enumerate(norm_locus_headers)
        if canon
    ]

    wells: Dict[str, WellLayout] = {}
    seen_loci: set[str] = set()
    r = header_row_idx + 1
    max_cols = len(norm_headers)

    while r <= ws.max_row:
        row_vals = [ws.cell(r, c).value for c in range(1, max_cols + 1)]
        raw_id = _clean(row_vals[id_col_idx]) if id_col_idx < len(row_vals) else None

        if not raw_id:
            break

        well_id = _normalize_well_id(raw_id)

        if not well_id:
            r += 1
            continue

        combo_id = (
            _clean(row_vals[combo_col_idx])
            if combo_col_idx is not None and combo_col_idx < len(row_vals)
            else None
        )
        race = (
            _clean(row_vals[race_col_idx])
            if race_col_idx is not None and race_col_idx < len(row_vals)
            else None
        )

        loci_dict: Dict[str, List[str]] = {}

        for idx, locus_label in locus_cols:
            v = _clean(row_vals[idx]) if idx < len(row_vals) else None

            if not v or v == "-":
                continue

            loci_dict.setdefault(locus_label, []).append(v)
            seen_loci.add(locus_label)

        wells[well_id] = WellLayout(
            well_id=well_id,
            combo_id=combo_id,
            race=race,
            loci=LociMap(data=loci_dict),
        )
        r += 1

    plate_fmt = _infer_plate_format_from_wells(list(wells.keys()))

    if plate_fmt not in COMMON_PLATE_SIZES:
        warnings.append(f"Inferred plate size '{plate_fmt}' is not a common plate size.")

    classic = set(ALLOWED_LOCI_CANON.values())
    custom_loci = [
        locus
        for locus in sorted(seen_loci)
        if locus not in classic
    ]

    layout = ParsedPlateLayout(
        upload_id="",
        schema_version="v1",
        sha256=sha256,
        lot_no=lot_no,
        compl_no=compl_no,
        plate_format=plate_fmt or DEFAULT_PLATE_FORMAT,
        wells={wid: wells[wid] for wid in sorted(wells.keys())},
        custom_loci=custom_loci,
        warnings=warnings,
        valid=True,
    )

    return sha256, layout
